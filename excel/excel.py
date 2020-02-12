import openpyxl
import re

class handlerErp(object):
    def __init__(self, workbook):
        self.wb     = openpyxl.load_workbook(workbook)
        self.sheet  = self.wb.active
        self.maxRow = self.sheet.max_row
        self.maxCol = self.sheet.max_column
        #ERP关键的型号信息在A,B列
        self.colA   = self.sheet['A']
        self.colB   = self.sheet['B']

    
    def getRowValues(self, row):
        rowData = []
        for i in range(1, self.maxCol + 1):
            cellValue = self.sheet.cell(row=row,column=i).value
            rowData.append(cellValue)
        return rowData

    def findout(self, oldname, newname):
        #create new workbook
        newWb = openpyxl.Workbook()
        newSheet = newWb.active
        oldKeyinfo = []
        newKeyinfo = []
        
        oldA       = []
        oldB       = []
        for i in range(len(self.colB)):
            s = str(self.colB[i].value)
            if re.search(oldname, s):
                oldKeyinfo.append(s)
                s_new = re.sub(oldname,newname,s)
                newKeyinfo.append(s_new)
                oldA.append(str(self.colA[i].value))
                oldB.append(str(self.colB[i].value))
                print('替换关系生成：{}--->{}'.format(s,s_new))

        for k in range(len(newKeyinfo)):
            for j in range(len(self.colB)):
                s = str(self.colB[j].value)
                if re.fullmatch(newKeyinfo[k],s):
                    row = self.getRowValues(j+1)
                    row.insert(0,oldA[k])
                    row.insert(1,oldB[k])
                    # row.insert(0,code)
                    print(row)
                    newSheet.append(row) 
        
                     
        newWb.save('替换结果.xlsx')     
        return oldKeyinfo,newKeyinfo    
        
if __name__ == '__main__':
    erp = handlerErp('./erp.xlsx') 
    
    erp.findout('ME909s-120','EC25-E')
    print('Hello,this is hanler function')
   